from io import BytesIO

from openpyxl import Workbook

from backend.models.tricount import Tricount
from backend.services.balance import compute_balances
from backend.services.settlement import compute_settlements


def export_tricount_to_excel(tricount: Tricount) -> BytesIO:
    wb = Workbook()

    ws_exp = wb.active
    ws_exp.title = "Dépenses"
    ws_exp.append(
       ["Description", "Montant", "Devise", "Payeur", "Participants", "Poids"]
    )

    for expense in tricount.expenses:
        payer_name = ""

        for user in tricount.users:
            if user.id == expense.payer_id:
                payer_name = user.name
                break

        participant_names = []

        for user in tricount.users:
            if user.id in expense.participants_ids:
                participant_names.append(user.name)
        
        weights_list = []

        for participant_id in expense.participants_ids:
            weight = expense.weights.get(participant_id, 1)
            weights_list.append(weight)

        ws_exp.append([
            expense.description,
            expense.amount,
            expense.currency.value,
            payer_name,
            ", ".join(participant_names),
            ", ".join(str(weight) for weight in weights_list),
        ])


    ws_bal = wb.create_sheet(title="Soldes")
    ws_bal.append(["Utilisateur", "Solde"])

    balances = compute_balances(tricount=tricount)
    for user in tricount.users:
        ws_bal.append([user.name, round(balances.get(user.id, 0.0), 2)])

    ws_set = wb.create_sheet(title="Règlements")
    ws_set.append(["De", "Vers", "Montant"])

    settlements = compute_settlements(balances=balances)
    for from_id, to_id, amount in settlements:
        from_name = ""
        to_name = ""

        for user in tricount.users:
            if user.id == from_id:
                from_name = user.name
            if user.id == to_id:
                to_name = user.name

        ws_set.append([from_name, to_name, amount])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
